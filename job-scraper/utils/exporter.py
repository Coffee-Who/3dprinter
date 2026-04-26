"""exporter.py - 匯出 Excel"""
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

COL_MAP = {
    "職缺名稱": "title",
    "公司名稱": "company",
    "工作地點": "location",
    "薪資":     "salary",
    "工作內容": "content",
    "技能需求": "skills",
    "學歷要求": "education",
    "工作年資": "experience",
    "來源":     "source",
    "連結":     "url",
    "抓取日期": "date",
}

COL_WIDTH = {
    "職缺名稱": 30,
    "公司名稱": 22,
    "工作地點": 14,
    "薪資":     16,
    "工作內容": 50,
    "技能需求": 35,
    "學歷要求": 14,
    "工作年資": 14,
    "來源":     10,
    "連結":     45,
    "抓取日期": 14,
}

# 顏色設定
HDR_BG   = "1F3864"   # 深藍
HDR_FONT = "FFFFFF"
ROW_104  = "FFF3E0"   # 淡橘
ROW_1111 = "E3F2FD"   # 淡藍
ROW_ALT  = "F5F5F5"
BORDER_C = "D0D0D0"

def export_excel(jobs: list, columns: list) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: 全部職缺 ──
    ws = wb.active
    ws.title = "全部職缺"
    _write_sheet(ws, jobs, columns)

    # ── Sheet 2: 僅 104 ──
    j104 = [j for j in jobs if j.get("source") == "104"]
    if j104:
        ws2 = wb.create_sheet("104人力銀行")
        _write_sheet(ws2, j104, columns)

    # ── Sheet 3: 僅 1111 ──
    j1111 = [j for j in jobs if j.get("source") == "1111"]
    if j1111:
        ws3 = wb.create_sheet("1111人力銀行")
        _write_sheet(ws3, j1111, columns)

    # ── Sheet 4: 統計摘要 ──
    ws_stat = wb.create_sheet("統計摘要")
    _write_summary(ws_stat, jobs)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_sheet(ws, jobs: list, columns: list):
    thin = Side(style="thin", color=BORDER_C)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 標題列
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = Font(bold=True, color=HDR_FONT, size=11)
        cell.fill      = PatternFill("solid", fgColor=HDR_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTH.get(col, 18)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # 資料列
    for ri, job in enumerate(jobs, 2):
        src = job.get("source", "")
        bg  = ROW_104 if src == "104" else (ROW_1111 if src == "1111" else
              (ROW_ALT if ri % 2 == 0 else "FFFFFF"))
        fill = PatternFill("solid", fgColor=bg)

        for ci, col in enumerate(columns, 1):
            key = COL_MAP.get(col, col)
            val = job.get(key, "")
            if isinstance(val, list):
                val = "、".join(val)

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font      = Font(size=10)

            # 連結超連結
            if col == "連結" and val and val.startswith("http"):
                cell.hyperlink = val
                cell.font      = Font(size=10, color="1F6FEB", underline="single")

            # 來源色
            if col == "來源":
                color = "D4521A" if val == "104" else "1565C0"
                cell.font = Font(size=10, bold=True, color=color)

        ws.row_dimensions[ri].height = 45

    # 自動篩選
    if jobs:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(columns))}{len(jobs)+1}"
        )


def _write_summary(ws, jobs: list):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18

    title_font  = Font(bold=True, size=14, color="1F3864")
    label_font  = Font(bold=True, size=11)
    value_font  = Font(size=11, color="1F6FEB")
    hdr_fill    = PatternFill("solid", fgColor="EBF1F7")

    ws["A1"] = "📊 職缺統計摘要"
    ws["A1"].font      = title_font
    ws["A2"] = f"產生時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font      = Font(size=10, color="888888")

    rows = [
        ("", ""),
        ("📋 總職缺數",       len(jobs)),
        ("🔶 104 人力銀行",   len([j for j in jobs if j.get("source")=="104"])),
        ("🔷 1111 人力銀行",  len([j for j in jobs if j.get("source")=="1111"])),
        ("✨ 今日新增",        len([j for j in jobs if j.get("date","")==datetime.now().strftime("%Y-%m-%d")])),
        ("", ""),
    ]

    # 薪資統計
    salaries = [j.get("salary_min",0) for j in jobs if j.get("salary_min",0) > 0]
    if salaries:
        rows += [
            ("💰 平均薪資下限",   f"{int(sum(salaries)/len(salaries)):,} 元"),
            ("💰 最高薪資下限",   f"{max(salaries):,} 元"),
        ]

    # 熱門技能
    from collections import Counter
    all_skills = [s for j in jobs for s in j.get("skills", [])]
    top_skills = Counter(all_skills).most_common(10)

    rows.append(("", ""))
    rows.append(("🔧 熱門技能 TOP 10", "出現次數"))

    r = len(rows) + 3
    for label, val in rows:
        ws.cell(row=r-len(rows)+1+rows.index((label,val)), column=1, value=label).font = label_font
        cell_v = ws.cell(row=r-len(rows)+1+rows.index((label,val)), column=2, value=val)
        cell_v.font = value_font

    base = len(rows) + 4
    for skill, cnt in top_skills:
        ws.cell(row=base, column=1, value=skill).font = Font(size=10)
        ws.cell(row=base, column=2, value=cnt).font   = Font(size=10, color="1F6FEB", bold=True)
        base += 1
