"""
Excel 輸出模組
支援：
  - 選項 A：每日新建一個 Excel 檔（檔名含日期）
  - 選項 B：統一寫入同一個 Excel，並自動過濾重複職缺
"""

from datetime import datetime
from pathlib import Path
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

COLUMNS = [
    "日期", "公司名稱", "職缺名稱", "薪資範圍",
    "工作地點", "職缺連結", "關鍵字", "關鍵字匹配段落", "來源",
]

COL_WIDTHS = {
    "日期": 14,
    "公司名稱": 22,
    "職缺名稱": 30,
    "薪資範圍": 20,
    "工作地點": 18,
    "職缺連結": 50,
    "關鍵字": 14,
    "關鍵字匹配段落": 55,
    "來源": 14,
}

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
ALT_FILL = PatternFill("solid", start_color="D6E4F0")
BORDER_SIDE = Side(style="thin", color="B0B0B0")
THIN_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE, bottom=BORDER_SIDE,
)


def _apply_header(ws):
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 20)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _write_row(ws, row_idx: int, job: dict):
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        value = job.get(col_name, "")
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="center", wrap_text=(col_name == "關鍵字匹配段落"))
        cell.border = THIN_BORDER
        # 超連結格式
        if col_name == "職缺連結" and value:
            cell.hyperlink = value
            cell.font = Font(color="0563C1", underline="single", name="Arial", size=10)
        else:
            cell.font = Font(name="Arial", size=10)
        # 斑馬紋
        if row_idx % 2 == 0:
            cell.fill = ALT_FILL
    ws.row_dimensions[row_idx].height = 22


def _dedup_key(job: dict) -> str:
    return f"{job.get('公司名稱', '')}|{job.get('職缺名稱', '')}|{job.get('職缺連結', '')}"


# ──────────────────────────────────────────────
# 選項 A：每天產生新 Excel
# ──────────────────────────────────────────────
def export_daily(jobs: list[dict], output_dir: str = "output") -> str:
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y%m%d")
    filename = Path(output_dir) / f"3D產業職缺_{today}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"職缺_{today}"
    _apply_header(ws)

    for row_idx, job in enumerate(jobs, start=2):
        _write_row(ws, row_idx, job)

    # 自動篩選
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)
    logger.info(f"[Export-A] 已儲存 {len(jobs)} 筆 → {filename}")
    return str(filename)


# ──────────────────────────────────────────────
# 選項 B：統一存入同一個 Excel，自動去重
# ──────────────────────────────────────────────
def export_cumulative(jobs: list[dict], filepath: str = "output/3D產業職缺_累計.xlsx") -> str:
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    fp = Path(filepath)

    existing_keys: set[str] = set()

    if fp.exists():
        wb = openpyxl.load_workbook(fp)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 取公司名稱(col 2)、職缺名稱(col 3)、連結(col 6)
            key = f"{row[1] or ''}|{row[2] or ''}|{row[5] or ''}"
            existing_keys.add(key)
        start_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "3D產業職缺"
        _apply_header(ws)
        start_row = 2

    new_jobs = [j for j in jobs if _dedup_key(j) not in existing_keys]
    for row_idx, job in enumerate(new_jobs, start=start_row):
        _write_row(ws, row_idx, job)

    ws.auto_filter.ref = ws.dimensions
    wb.save(fp)
    logger.info(f"[Export-B] 新增 {len(new_jobs)} 筆（跳過 {len(jobs)-len(new_jobs)} 重複）→ {fp}")
    return str(fp)
