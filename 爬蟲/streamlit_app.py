"""
3D 產業徵才爬蟲 - Streamlit Cloud 介面
"""

import io
import logging
from datetime import datetime

import streamlit as st
import pandas as pd

from crawler_104 import fetch_104_jobs, KEYWORDS
from crawler_1111 import fetch_1111_jobs
from exporter import export_daily, export_cumulative

# ── 頁面設定 ─────────────────────────────────
st.set_page_config(
    page_title="3D 產業職缺爬蟲",
    page_icon="🔍",
    layout="wide",
)

st.title("🔍 3D 產業徵才自動爬蟲")
st.caption("自動從 104、1111 人力銀行抓取近 24 小時的 3D 相關職缺")

# ── 側邊欄設定 ────────────────────────────────
with st.sidebar:
    st.header("⚙️ 設定")

    selected_keywords = st.multiselect(
        "搜尋關鍵字",
        options=KEYWORDS,
        default=KEYWORDS,
        help="可多選，留空則搜尋全部"
    )

    selected_sources = st.multiselect(
        "目標網站",
        options=["104人力銀行", "1111人力銀行"],
        default=["104人力銀行", "1111人力銀行"],
    )

    max_pages = st.slider("每個關鍵字最多翻頁數", 1, 10, 3)

    export_mode = st.radio(
        "輸出模式",
        ["每次產生新檔（含日期）", "累計到同一個檔案"],
        index=0,
    )

    st.divider()
    run_btn = st.button("🚀 開始爬取", use_container_width=True, type="primary")

# ── 主畫面 ────────────────────────────────────
if run_btn:
    if not selected_keywords:
        st.warning("請至少選擇一個關鍵字")
        st.stop()

    all_jobs = []
    progress = st.progress(0, text="準備中...")
    log_box = st.empty()
    logs = []

    total_tasks = len(selected_keywords) * len(selected_sources)
    done = 0

    for kw in selected_keywords:
        if "104人力銀行" in selected_sources:
            logs.append(f"🔎 104 | 關鍵字：{kw}")
            log_box.code("\n".join(logs))
            jobs = fetch_104_jobs(kw, max_pages=max_pages)
            all_jobs.extend(jobs)
            logs.append(f"   ✅ 找到 {len(jobs)} 筆")
            done += 1
            progress.progress(done / total_tasks, text=f"進度 {done}/{total_tasks}")

        if "1111人力銀行" in selected_sources:
            logs.append(f"🔎 1111 | 關鍵字：{kw}")
            log_box.code("\n".join(logs))
            jobs = fetch_1111_jobs(kw, max_pages=max_pages)
            all_jobs.extend(jobs)
            logs.append(f"   ✅ 找到 {len(jobs)} 筆")
            done += 1
            progress.progress(done / total_tasks, text=f"進度 {done}/{total_tasks}")

    progress.progress(1.0, text="爬取完成！")

    # 去重
    seen = set()
    deduped = []
    for j in all_jobs:
        key = f"{j.get('公司名稱','')}|{j.get('職缺名稱','')}"
        if key not in seen:
            seen.add(key)
            deduped.append(j)

    logs.append(f"\n📊 合計 {len(all_jobs)} 筆，去重後 {len(deduped)} 筆")
    log_box.code("\n".join(logs))

    if not deduped:
        st.info("🙁 過去 24 小時內沒有符合條件的新職缺")
        st.stop()

    # ── 顯示結果表格 ──
    st.success(f"✅ 共找到 **{len(deduped)}** 筆近 24 小時職缺")
    df = pd.DataFrame(deduped)

    # 讓連結可點擊
    st.dataframe(
        df,
        column_config={
            "職缺連結": st.column_config.LinkColumn("職缺連結"),
        },
        use_container_width=True,
        height=500,
    )

    # ── 統計圖表 ──
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("關鍵字分布")
        kw_counts = df["關鍵字"].value_counts()
        st.bar_chart(kw_counts)
    with col2:
        st.subheader("來源分布")
        src_counts = df["來源"].value_counts()
        st.bar_chart(src_counts)

    # ── 下載 Excel ──
    st.divider()
    st.subheader("📥 下載 Excel")

    # 產生 Excel bytes（in memory）
    output = io.BytesIO()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COLUMNS = ["日期","公司名稱","職缺名稱","薪資範圍","工作地點","職缺連結","關鍵字","關鍵字匹配段落","來源"]
    COL_WIDTHS = {"日期":14,"公司名稱":22,"職缺名稱":30,"薪資範圍":20,"工作地點":18,"職缺連結":50,"關鍵字":14,"關鍵字匹配段落":55,"來源":14}
    HEADER_FILL = PatternFill("solid", start_color="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    ALT_FILL = PatternFill("solid", start_color="D6E4F0")
    BD = Side(style="thin", color="B0B0B0")
    BORDER = Border(left=BD, right=BD, top=BD, bottom=BD)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3D職缺"

    for ci, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 20)
    ws.freeze_panes = "A2"

    for ri, job in enumerate(deduped, 2):
        for ci, col in enumerate(COLUMNS, 1):
            val = job.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=(col=="關鍵字匹配段落"))
            cell.border = BORDER
            if col == "職缺連結" and val:
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single", name="Arial", size=10)
            else:
                cell.font = Font(name="Arial", size=10)
            if ri % 2 == 0:
                cell.fill = ALT_FILL

    ws.auto_filter.ref = ws.dimensions
    wb.save(output)
    output.seek(0)

    today = datetime.now().strftime("%Y%m%d")
    filename = f"3D產業職缺_{today}.xlsx"

    st.download_button(
        label="⬇️ 下載 Excel 檔案",
        data=output,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )

else:
    st.info("👈 請在左側設定條件後，點擊「開始爬取」按鈕")

    with st.expander("📖 使用說明"):
        st.markdown("""
        1. 在左側選擇要搜尋的**關鍵字**與**目標網站**
        2. 調整**翻頁數**（越多越完整，但耗時較長）
        3. 點擊「開始爬取」按鈕
        4. 等待爬取完成後，可在畫面上預覽結果，並下載 Excel 檔案

        **搜尋關鍵字：** 3D掃描、3D列印、3D Scanning、3D Printing、逆向工程

        **注意：** Streamlit Cloud 免費版每次操作後如無人使用會進入休眠，重新點開需等待約 30 秒。
        """)
